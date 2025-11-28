import os
from pyopenms import MzMLFile, MSExperiment
from openpyxl import Workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import shutil  
def main(input_folder, split_folder):
    os.makedirs(split_folder, exist_ok=True)
    for file in os.listdir(input_folder):
        if file.endswith(".mzML"):
            file_path = os.path.join(input_folder, file)
            experiment = MSExperiment()
            MzMLFile().load(file_path, experiment)
            workbook = Workbook()
            extract_and_process(experiment, workbook)
            if "Sheet" in workbook.sheetnames:
                std = workbook["Sheet"]
                workbook.remove(std)
            split_workbook(workbook, os.path.join(split_folder, f"output_{file[:-5]}"))
def extract_and_process(experiment, workbook):
    j = 0
    for scan in experiment:
        if scan.getMSLevel() == 2:
            precursor = scan.getPrecursors()[0] if scan.getPrecursors() else None
            if precursor:
                precursor_mz = round(precursor.getMZ(), 4)
                precursor_rt = round(scan.getRT(), 4)
                sheet = workbook.create_sheet(title=f"mz_{j}_rt_{j}")
                sheet.append(["Precursor m/z", "Precursor RT", "MS2mz", "MS2i"])
                sheet.append([precursor_mz, precursor_rt, "", ""])
                for mz, i in zip(*scan.get_peaks()):
                    if i > 0:
                        mz = round(mz, 4)
                        i = round(i, 4)
                        sheet.append(["", "", mz, i])
                j += 1
def split_workbook(workbook, output_prefix):
    output_dir = os.path.dirname(output_prefix)
    if output_dir:  
        os.makedirs(output_dir, exist_ok=True)
    sheet_names = workbook.sheetnames
    batch_size = 2000
    current_workbook_index = 1
    current_sheet_count = 0
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        output_sheet = output_workbook.create_sheet(title=sheet_name)
        copy_sheet(sheet, output_sheet)
        current_sheet_count += 1
        if current_sheet_count == batch_size:
            output_file = f"{output_prefix}_{current_workbook_index}.xlsx"
            output_workbook.save(output_file)
            output_workbook = Workbook()
            output_workbook.remove(output_workbook.active)
            current_workbook_index += 1
            current_sheet_count = 0
    if current_sheet_count > 0:
        output_file = f"{output_prefix}_{current_workbook_index}.xlsx"
        output_workbook.save(output_file)
def copy_sheet(source_sheet, target_sheet):
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)
def classify_excel(input_filename, output_filename):
    xls = pd.ExcelFile(input_filename)
    with pd.ExcelWriter(output_filename, engine='xlsxwriter') as writer:
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'MS2mz' not in df.columns:
                    continue
                df['CLASS'] = None
                class_values = df['CLASS']
                MS2mz_values = df['MS2mz']
                if (((MS2mz_values - 184.073).abs() <= 0.5).any() and
                        (((MS2mz_values - 224.107).abs() <= 0.5).any() or
                         ((MS2mz_values - 226.085).abs() <= 0.5).any()) and
                        not ((MS2mz_values - 142.026).abs() <= 0.5).any()):
                    pc_index = class_values.index[class_values.isnull()].min()
                    df.loc[pc_index, 'CLASS'] = 'PC'
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                elif (((MS2mz_values - 184.073).abs() <= 0.5).any() and
                      (((MS2mz_values - 224.107).abs() <= 0.5).any() or
                       ((MS2mz_values - 226.085).abs() <= 0.5).any()) and
                      ((MS2mz_values - 142.026).abs() <= 0.5).any() and
                      ((MS2mz_values - 182.058).abs() <= 0.5).any()):
                    df_pc = df.copy()
                    df_pe = df.copy()
                    pc_index = class_values.index[class_values.isnull()].min()
                    df_pc.loc[pc_index, 'CLASS'] = 'PC'
                    df_pe.loc[pc_index, 'CLASS'] = 'PE'
                    df_pc.to_excel(writer, sheet_name=sheet_name + '_PC', index=False)
                    df_pe.to_excel(writer, sheet_name=sheet_name + '_PE', index=False)
                elif (((MS2mz_values - 184.073).abs() <= 0.5).any() and
                      (((MS2mz_values - 225.100).abs() <= 0.5).any() or
                       ((MS2mz_values - 253.085).abs() <= 0.5).any())):
                    sm_index = class_values.index[class_values.isnull()].min()
                    df.loc[sm_index, 'CLASS'] = 'SM'
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                elif (((MS2mz_values - 142.026).abs() <= 0.5).any() and
                      ((MS2mz_values - 182.058).abs() <= 0.5).any()):
                    pe_index = class_values.index[class_values.isnull()].min()
                    df.loc[pe_index, 'CLASS'] = 'PE'
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    uncertain_index = class_values.index[class_values.isnull()].min()
                    df.loc[uncertain_index, 'CLASS'] = 'Uncertain'
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                return 0
def process_files_and_profile(input_folder, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    for file_name in os.listdir(input_folder):
        if file_name.endswith(".xlsx"):
            input_file_path = os.path.join(input_folder, file_name)
            output_file_path = os.path.join(output_folder, file_name)
            classify_excel(input_file_path, output_file_path)
def process_uncertain(df, idx_uncertain):
    mz0 = df.at[0, 'Precursor m/z']
    diffs = (idx_uncertain['TheoMz'] - mz0).abs()
    i = diffs.idxmin()
    if diffs[i] >= 0.01:
        return None
    matched = idx_uncertain.loc[i]
    df.loc[0, matched.index] = matched.values
    return df

def process_others(df, idx_other):
    mz0 = df.at[0, 'Precursor m/z']
    cls = df.at[0, 'CLASS']
    sub = idx_other[idx_other['MAIN_CLASS'] == cls]
    if sub.empty:
        return None
    diffs = (sub['TheoMz'] - mz0).abs()
    min_diff = diffs.min()
    if min_diff >= 0.01:
        return None
    hits = sub[diffs == min_diff]
    sheets = []
    for _, row in hits.iterrows():
        new = df.copy()
        new.loc[0, row.index] = row.values
        sheets.append(new)
    return sheets[0] if len(sheets) == 1 else sheets

def process_single_file(fn, idx_uncertain, idx_other, out_folder, verbose=False):
    if verbose:
        print(f"→ Processing {os.path.basename(fn)}")
    xls = pd.ExcelFile(fn, engine='openpyxl')
    outputs = []

    for name in xls.sheet_names:
        if verbose:
            print(f"   • {name}", end="")
        df = pd.read_excel(xls, sheet_name=name, engine='openpyxl')

        if 'CLASS' not in df.columns:
            if verbose:
                print("  [skipped: no CLASS]")
            continue

        if df.at[0, 'CLASS'] == 'Uncertain':
            res = process_uncertain(df, idx_uncertain)
        else:
            res = process_others(df, idx_other)

        if res is None:
            if verbose:
                print("  [no match]")
        else:
            if isinstance(res, list):
                for i, sheet in enumerate(res, 1):
                    outputs.append((f"{name}_{i}", sheet))
                if verbose:
                    print(f"  [found {len(res)}]")
            else:
                outputs.append((name, res))
                if verbose:
                    print("  [✓]")

    if outputs:
        out_path = os.path.join(out_folder, os.path.basename(fn))
        wb = Workbook()
        wb.remove(wb.active)
        for nm, sheet in outputs:
            ws = wb.create_sheet(title=nm)
            for row in dataframe_to_rows(sheet, index=False, header=True):
                ws.append(row)
        wb.save(out_path)
        if verbose:
            print(f"   → saved to {out_path}")
    else:
        if verbose:
            print("   → nothing to save")

def process_all(input_folder, index_file, output_folder, verbose=False):
    idx_unc = pd.read_excel(index_file, sheet_name='Uncertain', engine='openpyxl')
    idx_oth = pd.read_excel(index_file, sheet_name='Others',    engine='openpyxl')
    os.makedirs(output_folder, exist_ok=True)

    files = sorted(
        os.path.join(input_folder, f)
        for f in os.listdir(input_folder)
        if f.lower().endswith('.xlsx')
    )
    for fn in files:
        process_single_file(fn, idx_unc, idx_oth, output_folder, verbose=verbose)

def process_PC(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-C', 'sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
        input_data[f'{sn}-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-C-match'] = index_data.at[j, 'sn-2-C diagnostic ions']
                input_data.at[i, 'sn-2-C-chain'] = index_data.at[j, 'sn-2-C-chain']
                input_data.at[i, 'sn-2-C-MS2i'] = input_data.at[i, 'MS2i']
                sn_2_diagnostic_ion = index_data.at[j, 'sn-2 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-match'] = sn_2_diagnostic_ion
                        input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                        input_data.at[i, 'sn-2-MS2i'] = input_data.at[k, 'MS2i']
                        if input_data.at[i, 'sn-2-C-MS2i'] < input_data.at[i, 'sn-2-MS2i']:
                            input_data.at[i, 'sn-2-C-match'] = ''
                            input_data.at[i, 'sn-2-C-chain'] = ''
                            input_data.at[i, 'sn-2-C-MS2i'] = ''
                            input_data.at[i, 'sn-2-match'] = ''
                            input_data.at[i, 'sn-2-chain'] = ''
                            input_data.at[i, 'sn-2-MS2i'] = ''
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    # definite the format of Composition column as x:y
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    # recording the first data of 'Composition' column
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    # definite the sn-2-O-chain as “a:b”; sn-1-chain as “c:d”;
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PC({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['sn-2-MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        input_data.at[index, 'sn-2-C-diagnostic ions'] = row['sn-2-C-match']
                        input_data.at[index, 'sn-2-C-intensity'] = row['sn-2-C-MS2i']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_PCO(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-C', 'sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
        input_data[f'{sn}-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-C-match'] = index_data.at[j, 'sn-2-C diagnostic ions']
                input_data.at[i, 'sn-2-C-chain'] = index_data.at[j, 'sn-2-C-chain']
                sn_2_o_diagnostic_ion = index_data.at[j, 'sn-2 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-match'] = sn_2_o_diagnostic_ion
                        input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                        input_data.at[i, 'sn-2-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PC-O({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['sn-2-MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_PCP(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-C', 'sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
        input_data[f'{sn}-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-C-match'] = index_data.at[j, 'sn-2-C diagnostic ions']
                input_data.at[i, 'sn-2-C-chain'] = index_data.at[j, 'sn-2-C-chain']
                sn_2_o_diagnostic_ion = index_data.at[j, 'sn-2 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-match'] = sn_2_o_diagnostic_ion
                        input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                        input_data.at[i, 'sn-2-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PC-P({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_PE(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-C', 'sn-2', 'sn-1-C', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
        input_data['sn-2-MS2i'] = ''
        input_data['sn-1-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-C-match'] = index_data.at[j, 'sn-2-C diagnostic ions']
                input_data.at[i, 'sn-2-C-chain'] = index_data.at[j, 'sn-2-C-chain']
                sn_2_o_diagnostic_ion = index_data.at[j, 'sn-2 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-match'] = sn_2_o_diagnostic_ion
                        input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                        input_data.at[i, 'sn-2-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-C-match'] = index_data.at[j, 'sn-1-C diagnostic ions']
                input_data.at[i, 'sn-1-C-chain'] = index_data.at[j, 'sn-1-C-chain']
                sn_1_o_diagnostic_ion = index_data.at[j, 'sn-1 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_1_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-1-match'] = sn_1_o_diagnostic_ion
                        input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
                        input_data.at[i, 'sn-1-MS2i'] = input_data.at[k, 'MS2i']
                        break
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PE({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['sn-1-MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['sn-2-MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_PEO(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-match'] = index_data.at[j, 'sn-2 diagnostic ions']
                input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PE-O({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_PEP(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-match'] = index_data.at[j, 'sn-2 diagnostic ions']
                input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"PE-P({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_SM(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-1-O', 'sn-1', 'sn-2']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
    input_data['sn-1-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1-O diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-O-match'] = index_data.at[j, 'sn-1-O diagnostic ions']
                input_data.at[i, 'sn-1-O-chain'] = index_data.at[j, 'sn-1-O-chain']
                sn_1_o_diagnostic_ion = index_data.at[j, 'sn-1 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_1_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-1-match'] = sn_1_o_diagnostic_ion
                        input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
                        input_data.at[i, 'sn-1-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-match'] = index_data.at[j, 'sn-2 diagnostic ions']
                input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"SM(d{row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['sn-1-MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_Cer(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-O', 'sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
    input_data['sn-2-O-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-match'] = index_data.at[j, 'sn-2 diagnostic ions']
                input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                sn_2_o_diagnostic_ion = index_data.at[j, 'sn-2-O diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-O-match'] = sn_2_o_diagnostic_ion
                        input_data.at[i, 'sn-2-O-chain'] = index_data.at[j, 'sn-2-O-chain']
                        input_data.at[i, 'sn-2-O-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"Cer(d{row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_Cer2OH(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2 diagnostic ion']) <= 0.05:
                input_data.at[i, 'sn-2-match'] = index_data.at[j, 'sn-2 diagnostic ion']
                input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ion']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ion']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"Cer(d{row2['sn-1-chain']}/{row['sn-2-chain']}(2OH))"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_DG(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-2-C', 'sn-2', 'sn-1']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chain'] = ''
        input_data[f'{sn}-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-2-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-2-C-match'] = index_data.at[j, 'sn-2-C diagnostic ions']
                input_data.at[i, 'sn-2-C-chain'] = index_data.at[j, 'sn-2-C-chain']
                input_data.at[i, 'sn-2-C-MS2i'] = input_data.at[i, 'MS2i']
                sn_2_c_diagnostic_ion = index_data.at[j, 'sn-2 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_2_c_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-2-match'] = sn_2_c_diagnostic_ion
                        input_data.at[i, 'sn-2-chain'] = index_data.at[j, 'sn-2-chain']
                        input_data.at[i, 'sn-2-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1 diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-match'] = index_data.at[j, 'sn-1 diagnostic ions']
                input_data.at[i, 'sn-1-chain'] = index_data.at[j, 'sn-1-chain']
                input_data.at[i, 'sn-1-MS2i'] = input_data.at[i, 'MS2i']
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chain'] = input_data['sn-2-chain'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chain'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chain'] = input_data['sn-1-chain'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chain'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chain'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    if sn_2_chain_a + sn_1_chain_c == first_composition_x and sn_2_chain_b + sn_1_chain_d == first_composition_y:
                        input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                        input_data.at[index, 'sn-composition'] = f"DG({row2['sn-1-chain']}/{row['sn-2-chain']})"
                        input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                        input_data.at[index, 'sn-1-intensity'] = row2['sn-1-MS2i']
                        input_data.at[index, 'sn-1-com'] = row2['sn-1-chain']
                        input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                        input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                        input_data.at[index, 'sn-2-intensity'] = row['sn-2-MS2i']
                        input_data.at[index, 'sn-2-com'] = row['sn-2-chain']
                        input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                        input_data.at[index, 'sn-2-C-diagnostic ions'] = row['sn-2-C-match']
                        input_data.at[index, 'sn-2-C-intensity'] = row['sn-2-C-MS2i']
                        break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_TG(input_data, index_file, index_sheet_name, columns_to_remove):
    index_data = pd.read_excel(index_file, sheet_name=index_sheet_name)
    for sn in ['sn-1-C', 'sn-1', 'sn-3-C', 'sn-3', 'sn-2']:
        input_data[f'{sn}-match'] = ''
        input_data[f'{sn}-chains'] = ''
        input_data[f'{sn}-MS2i'] = ''
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-1-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-1-C-match'] = index_data.at[j, 'sn-1-C diagnostic ions']
                input_data.at[i, 'sn-1-C-chains'] = index_data.at[j, 'sn-1-C-chains']
                input_data.at[i, 'sn-1-C-MS2i'] = input_data.at[i, 'MS2i']
                sn_1_o_diagnostic_ion = index_data.at[j, 'sn-1 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_1_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-1-match'] = sn_1_o_diagnostic_ion
                        input_data.at[i, 'sn-1-chains'] = index_data.at[j, 'sn-1-chains']
                        input_data.at[i, 'sn-1-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)):
        for j in range(len(index_data)):
            if abs(input_data.at[i, 'MS2mz'] - index_data.at[j, 'sn-3-C diagnostic ions']) <= 0.05:
                input_data.at[i, 'sn-3-C-match'] = index_data.at[j, 'sn-3-C diagnostic ions']
                input_data.at[i, 'sn-3-C-chains'] = index_data.at[j, 'sn-3-C-chains']
                input_data.at[i, 'sn-3-C-MS2i'] = input_data.at[i, 'MS2i']
                sn_3_o_diagnostic_ion = index_data.at[j, 'sn-3 diagnostic ions']
                for k in range(len(input_data)):
                    if abs(input_data.at[k, 'MS2mz'] - sn_3_o_diagnostic_ion) <= 0.05:
                        input_data.at[i, 'sn-3-match'] = sn_3_o_diagnostic_ion
                        input_data.at[i, 'sn-3-chains'] = index_data.at[j, 'sn-3-chains']
                        input_data.at[i, 'sn-3-MS2i'] = input_data.at[k, 'MS2i']
                        break
    for i in range(len(input_data)): #sn-2 only output the best mathed 'MS2mz';
        ms2mz = input_data.at[i, 'MS2mz']
        min_difference = float('inf')
        min_index = -1
        for j in range(len(index_data)):
            difference = abs(ms2mz - index_data.at[j, 'sn-2 diagnostic ions'])
            if difference <= 0.01 and (input_data.at[i, 'sn-2-match'] is None or difference < min_difference):
                min_difference = difference
                min_index = j
        if min_index != -1:
            input_data.at[i, 'sn-2-match'] = index_data.at[min_index, 'sn-2 diagnostic ions']
            input_data.at[i, 'sn-2-chains'] = index_data.at[min_index, 'sn-2-chains']
            input_data.at[i, 'sn-2-MS2i'] = input_data.at[i, 'MS2i']
            # Clear previous matches if a new minimum difference is found
            for k in range(len(input_data)):
                if k != i and abs(input_data.at[k, 'MS2mz'] - index_data.at[min_index, 'sn-2 diagnostic ions']) <= 0.05:
                    input_data.at[k, 'sn-2-match'] = None
                    input_data.at[k, 'sn-2-chains'] = None
                    input_data.at[k, 'sn-2-MS2i'] = None
    input_data['x'] = input_data['Composition'].apply(lambda x: int(x.split(':')[0]) if not pd.isnull(x) else np.nan)
    input_data['y'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    first_composition_x, first_composition_y = None, None
    for index, row in input_data.iterrows():
        if not pd.isnull(row['Composition']):
            first_composition_x, first_composition_y = map(int, row['Composition'].split(':'))
            break
    input_data['sn-2-chains'] = input_data['sn-2-chains'].astype(str)
    input_data[['sn-2-chain-a', 'sn-2-chain-b']] = input_data['sn-2-chains'].str.extract(r'(\d+):(\d+)')
    input_data['sn-2-chains'] = input_data['sn-2-chain-a'] + ':' + input_data['sn-2-chain-b']
    input_data['sn-1-chains'] = input_data['sn-1-chains'].astype(str)
    input_data[['sn-1-chain-c', 'sn-1-chain-d']] = input_data['sn-1-chains'].str.extract(r'(\d+):(\d+)')
    input_data['sn-1-chains'] = input_data['sn-1-chain-c'] + ':' + input_data['sn-1-chain-d']
    input_data['sn-3-chains'] = input_data['sn-3-chains'].astype(str)
    input_data[['sn-3-chain-e', 'sn-3-chain-f']] = input_data['sn-3-chains'].str.extract(r'(\d+):(\d+)')
    input_data['sn-3-chains'] = input_data['sn-3-chain-e'] + ':' + input_data['sn-3-chain-f']
    input_data['sn-composition'] = np.nan
    for index, row in input_data.iterrows():
        if not pd.isnull(row['sn-2-chain-a']) and not pd.isnull(row['sn-2-chain-b']):
            sn_2_chain_a = int(row['sn-2-chain-a'])
            sn_2_chain_b = int(row['sn-2-chain-b'])
            for index2, row2 in input_data.iterrows():
                if not pd.isnull(row2['sn-1-chain-c']) and not pd.isnull(row2['sn-1-chain-d']):
                    sn_1_chain_c = int(row2['sn-1-chain-c'])
                    sn_1_chain_d = int(row2['sn-1-chain-d'])
                    for index3, row3 in input_data.iterrows():
                        if not pd.isnull(row3['sn-3-chain-e']) and not pd.isnull(row3['sn-3-chain-f']):
                            sn_3_chain_e = int(row3['sn-3-chain-e'])
                            sn_3_chain_f = int(row3['sn-3-chain-f'])
                            if sn_2_chain_a + sn_1_chain_c + sn_3_chain_e == first_composition_x and sn_2_chain_b + sn_1_chain_d + sn_3_chain_f == first_composition_y:
                                input_data['sn-composition'] = input_data['sn-composition'].astype(str)
                                input_data.at[index, 'sn-composition'] = f"TG({row2['sn-1-chains']}_/{row['sn-2-chains']}/_{row3['sn-3-chains']})"
                                input_data.at[index, 'sn-1-diagnostic ions'] = row2['sn-1-match']
                                input_data.at[index, 'sn-1-intensity'] = row2['sn-1-MS2i']
                                input_data.at[index, 'sn-1-com'] = row2['sn-1-chains']
                                input_data.at[index, 'sn_1_chain_d'] = row2['sn-1-chain-d']
                                input_data.at[index, 'sn-1-C-diagnostic ions'] = row2['sn-1-C-match']
                                input_data.at[index, 'sn-1-C-intensity'] = row2['sn-1-C-MS2i']
                                input_data.at[index, 'sn-2-diagnostic ions'] = row['sn-2-match']
                                input_data.at[index, 'sn-2-intensity'] = row['sn-2-MS2i']
                                input_data.at[index, 'sn-2-com'] = row['sn-2-chains']
                                input_data.at[index, 'sn_2_chain_b'] = row['sn-2-chain-b']
                                input_data.at[index, 'sn-3-diagnostic ions'] = row3['sn-3-match']
                                input_data.at[index, 'sn-3-intensity'] = row3['sn-3-MS2i']
                                input_data.at[index, 'sn-3-com'] = row3['sn-3-chains']
                                input_data.at[index, 'sn_3_chain_f'] = row3['sn-3-chain-f']
                                input_data.at[index, 'sn-3-C-diagnostic ions'] = row3['sn-3-C-match']
                                input_data.at[index, 'sn-3-C-intensity'] = row3['sn-3-C-MS2i']
                                break
    input_data = input_data.drop(columns=columns_to_remove, errors='ignore')
    return input_data
def process_LPC(input_data):
    input_data['sn_1_chain_d'] = input_data['Composition'].apply(lambda x: int(x.split(':')[1]) if not pd.isnull(x) else np.nan)
    return input_data
def process_original_data(input_folder, index_file, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    index_df = pd.read_excel(index_file)
    for filename in os.listdir(input_folder):
        input_file = os.path.join(input_folder, filename)
        output_file = os.path.join(output_folder, filename)
        xls = pd.ExcelFile(input_file)
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        for sheet_name in xls.sheet_names:
            input_data = pd.read_excel(input_file, sheet_name=sheet_name)
            processed_data = None
            columns_to_remove = None
            if 'Subclass' in input_data.columns:
                subclass_value = input_data['Subclass'].iloc[0]
                if subclass_value == 'unmatched':
                    continue
                if subclass_value == 'PC':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-MS2i', 'sn-2-C-MS2i', 'sn-1-MS2i', 'sn-2-chain-b', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-C-match', 'sn-2-C-chain', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PC(input_data, index_file, 'PC', columns_to_remove)
                elif subclass_value == 'PC-O':
                    columns_to_remove = ['x', 'y', 'sn-1-MS2i','sn-2-C-MS2i','sn-2-chain-a', 'sn-2-chain-b', 'sn-2-MS2i', 'sn-1-chain-c', 'sn-2-C-chain', 'sn-1-chain-d', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PCO(input_data, index_file, 'PC-O', columns_to_remove)
                elif subclass_value == 'PC-P':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-2-MS2i', 'sn-1-chain-c', 'sn-2-C-chain', 'sn-1-chain-d', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PCP(input_data, index_file, 'PC-P', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'PE':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-MS2i', 'sn-2-MS2i', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-C-chain', 'sn-1-C-match', 'sn-1-C-chain', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PE(input_data, index_file, 'PE', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'PE-O':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PEO(input_data, index_file, 'PE-O', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'PE-P':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_PEP(input_data, index_file, 'PE-P', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'SM':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-MS2i', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-1-O-match', 'sn-1-O-chain', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_SM(input_data, index_file, 'SM', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'Cer':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-2-O-MS2i', 'sn-2-O-match', 'sn-2-O-chain', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_Cer(input_data, index_file, 'Cer', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'Cer-2OH':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_Cer2OH(input_data, index_file, 'Cer-2OH', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'DG':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-2-C-MS2i', 'sn-2-C-chain', 'sn-2-MS2i', 'sn-1-MS2i', 'sn-1-chain-c', 'sn-1-O-match', 'sn-1-O-chain',
                                         'sn-1-chain-d', 'sn-2-C-match', 'sn-2-match', 'sn-2-chain', 'sn-1-match', 'sn-1-chain']
                    processed_data = process_DG(input_data, index_file, 'DG', columns_to_remove)
                elif input_data['Subclass'].iloc[0] == 'TG':
                    columns_to_remove = ['x', 'y', 'sn-2-chain-a', 'sn-2-chain-b', 'sn-1-C-match', 'sn-3-C-match', 'sn-1-chain-c', 'sn-1-chain-d', 'sn-3-match', 'sn-3-chain-e', 'sn-3-chain-f',
                                         'sn-2-C-match', 'sn-2-match', 'sn-1-match', 'sn-1-C-chains', 'sn-1-C-MS2i', 'sn-1-chains', 'sn-1-MS2i', 'sn-3-C-chains', 'sn-3-C-MS2i', 'sn-3-chains', 'sn-3-MS2i', 'sn-2-chains', 'sn-2-MS2i']
                    processed_data = process_TG(input_data, index_file, 'TG', columns_to_remove)
                elif input_data['Subclass'].iloc[0] in ['LPC', 'LPC-O', 'LPC-P', 'LPE', 'LPE-O', 'LPE-P']:
                    processed_data = process_LPC(input_data)
                else:
                    print(f"No processing function defined for Subclass: {input_data['Subclass'].iloc[0]}")
            if processed_data is not None:
                processed_data.loc[1:, ['Precursor m/z', 'Precursor RT', 'CLASS', 'MAIN_CLASS', 'Subclass', 'TheoMz', 'ID', 'Composition', 'Formula', 'Mass', 'Adduct ion']] = float('nan')
                processed_data.to_excel(writer, sheet_name=sheet_name, index=False)
        if len(writer.sheets) > 0:
            for sheet_name in writer.sheets.keys():
                writer.book[sheet_name].sheet_state = 'visible'
            writer.close()
        else:
            print(f"No processing function defined for Subclass: {subclass_value}")

def process_excel_DU1(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-1'] = first_precursor_mz - index_data['NL-1']
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NL1 = row['Precursor m/z-NL-1']
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        matched_input_data_NL1 = input_data[(input_data['MS2mz'] - ms2mz_NL1).abs() < 0.05]
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        if not matched_input_data_NL1.empty and not matched_input_data_NL2.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl1_intensity = ms2i_values[matched_input_data_NL1.index]
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl1_index = nl1_intensity.idxmax()
                max_nl2_index = nl2_intensity.idxmax()
                nl1_intensity = nl1_intensity[max_nl1_index]
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl1_intensity = ms2i_values[matched_input_data_NL1.index]
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl1_index = nl1_intensity.idxmax()
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl1_intensity = nl1_intensity[max_nl1_index]
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl1_intensity > nl_db_intensity).all() and (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL1.index, 'NL-1 Matched'] = row['Precursor m/z-NL-1']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
    columns_to_check = ['NL-1 Matched', 'NL-2 Matched', 'double bond-NL-2']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['NL-1 Matched'] = input_data['NL-1 Matched'].mask(input_data['NL-1 Matched'].duplicated(), '')
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
    return input_data

def process_excel_DU2(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-1'] = first_precursor_mz - index_data['NL-1']
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-4'] = first_precursor_mz - index_data['NL-4']
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NL1 = row['Precursor m/z-NL-1']
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NL4 = row['Precursor m/z-NL-4']
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        matched_input_data_NL1 = input_data[(input_data['MS2mz'] - ms2mz_NL1).abs() < 0.05]
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NL4 = input_data[(input_data['MS2mz'] - ms2mz_NL4).abs() < 0.05]
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        if not matched_input_data_NL1.empty and not matched_input_data_NL2.empty and not matched_input_data_NL4.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl1_intensity = ms2i_values[matched_input_data_NL1.index]
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl1_index = nl1_intensity.idxmax()
                max_nl2_index = nl2_intensity.idxmax()
                nl1_intensity = nl1_intensity[max_nl1_index]
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl1_intensity = ms2i_values[matched_input_data_NL1.index]
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl1_index = nl1_intensity.idxmax()
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl1_intensity = nl1_intensity[max_nl1_index]
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl1_intensity > nl_db_intensity).all() and (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL4.index, 'NL-4 Matched'] = row['Precursor m/z-NL-4']
    columns_to_check = ['double bond-NL-2', 'NL-2 Matched', 'NL-4 Matched']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
            input_data['NL-4 Matched'] = input_data['NL-4 Matched'].mask(input_data['NL-4 Matched'].duplicated(), '')
    return input_data
def process_excel_DU3(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-4'] = first_precursor_mz - index_data['NL-4']
    index_data['Precursor m/z-NL-6'] = first_precursor_mz - index_data['NL-6']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NL4 = row['Precursor m/z-NL-4']
        ms2mz_NL6 = row['Precursor m/z-NL-6']
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NL4 = input_data[(input_data['MS2mz'] - ms2mz_NL4).abs() < 0.05]
        matched_input_data_NL6 = input_data[(input_data['MS2mz'] - ms2mz_NL6).abs() < 0.05]
        if not matched_input_data_NL2.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl2_index = nl2_intensity.idxmax()
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL4.index, 'NL-4 Matched'] = row['Precursor m/z-NL-4']
                input_data.loc[matched_input_data_NL6.index, 'NL-6 Matched'] = row['Precursor m/z-NL-6']
    columns_to_check = ['double bond-NL-2', 'NL-2 Matched', 'NL-4 Matched', 'NL-6 Matched']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['NL-4 Matched'] = input_data['NL-4 Matched'].mask(input_data['NL-4 Matched'].duplicated(), '')
            input_data['NL-6 Matched'] = input_data['NL-6 Matched'].mask(input_data['NL-6 Matched'].duplicated(), '')
    return input_data
def process_excel_DU4(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-4'] = first_precursor_mz - index_data['NL-4']
    index_data['Precursor m/z-NL-6'] = first_precursor_mz - index_data['NL-6']
    index_data['Precursor m/z-NL-8'] = first_precursor_mz - index_data['NL-8']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NL4 = row['Precursor m/z-NL-4']
        ms2mz_NL6 = row['Precursor m/z-NL-6']
        ms2mz_NL8 = row['Precursor m/z-NL-8']
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NL4 = input_data[(input_data['MS2mz'] - ms2mz_NL4).abs() < 0.05]
        matched_input_data_NL6 = input_data[(input_data['MS2mz'] - ms2mz_NL6).abs() < 0.05]
        matched_input_data_NL8 = input_data[(input_data['MS2mz'] - ms2mz_NL8).abs() < 0.05]
        if not matched_input_data_NL2.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl2_index = nl2_intensity.idxmax()
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL4.index, 'NL-4 Matched'] = row['Precursor m/z-NL-4']
                input_data.loc[matched_input_data_NL6.index, 'NL-6 Matched'] = row['Precursor m/z-NL-6']
                input_data.loc[matched_input_data_NL8.index, 'NL-8 Matched'] = row['Precursor m/z-NL-8']
    columns_to_check = ['double bond-NL-2', 'NL-2 Matched', 'NL-4 Matched', 'NL-6 Matched', 'NL-8 Matched']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['NL-4 Matched'] = input_data['NL-4 Matched'].mask(input_data['NL-4 Matched'].duplicated(), '')
            input_data['NL-6 Matched'] = input_data['NL-6 Matched'].mask(input_data['NL-6 Matched'].duplicated(), '')
            input_data['NL-8 Matched'] = input_data['NL-8 Matched'].mask(input_data['NL-8 Matched'].duplicated(), '')
    return input_data
def process_excel_DU5(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-4'] = first_precursor_mz - index_data['NL-4']
    index_data['Precursor m/z-NL-6'] = first_precursor_mz - index_data['NL-6']
    index_data['Precursor m/z-NL-8'] = first_precursor_mz - index_data['NL-8']
    index_data['Precursor m/z-NL-10'] = first_precursor_mz - index_data['NL-10']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NL4 = row['Precursor m/z-NL-4']
        ms2mz_NL6 = row['Precursor m/z-NL-6']
        ms2mz_NL8 = row['Precursor m/z-NL-8']
        ms2mz_NL10 = row['Precursor m/z-NL-10']
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NL4 = input_data[(input_data['MS2mz'] - ms2mz_NL4).abs() < 0.05]
        matched_input_data_NL6 = input_data[(input_data['MS2mz'] - ms2mz_NL6).abs() < 0.05]
        matched_input_data_NL8 = input_data[(input_data['MS2mz'] - ms2mz_NL8).abs() < 0.05]
        matched_input_data_NL10 = input_data[(input_data['MS2mz'] - ms2mz_NL10).abs() < 0.05]
        if not matched_input_data_NL2.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl2_index = nl2_intensity.idxmax()
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL4.index, 'NL-4 Matched'] = row['Precursor m/z-NL-4']
                input_data.loc[matched_input_data_NL6.index, 'NL-6 Matched'] = row['Precursor m/z-NL-6']
                input_data.loc[matched_input_data_NL8.index, 'NL-8 Matched'] = row['Precursor m/z-NL-8']
                input_data.loc[matched_input_data_NL10.index, 'NL-10 Matched'] = row['Precursor m/z-NL-10']
    columns_to_check = ['double bond-NL-2', 'NL-2 Matched', 'NL-4 Matched', 'NL-6 Matched', 'NL-8 Matched', 'NL-10 Matched']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['NL-4 Matched'] = input_data['NL-4 Matched'].mask(input_data['NL-4 Matched'].duplicated(), '')
            input_data['NL-6 Matched'] = input_data['NL-6 Matched'].mask(input_data['NL-6 Matched'].duplicated(), '')
            input_data['NL-8 Matched'] = input_data['NL-8 Matched'].mask(input_data['NL-8 Matched'].duplicated(), '')
            input_data['NL-10 Matched'] = input_data['NL-10 Matched'].mask(input_data['NL-10 Matched'].duplicated(), '')
    return input_data
def process_excel_DU6(input_data, index_data):
    first_precursor_mz = input_data['Precursor m/z'].iloc[0]
    index_data['Precursor m/z-NL-2'] = first_precursor_mz - index_data['NL-2']
    index_data['Precursor m/z-NL-4'] = first_precursor_mz - index_data['NL-4']
    index_data['Precursor m/z-NL-6'] = first_precursor_mz - index_data['NL-6']
    index_data['Precursor m/z-NL-8'] = first_precursor_mz - index_data['NL-8']
    index_data['Precursor m/z-NL-10'] = first_precursor_mz - index_data['NL-10']
    index_data['Precursor m/z-NL-12'] = first_precursor_mz - index_data['NL-12']
    index_data['Precursor m/z-NL-db-1'] = first_precursor_mz - index_data['NL-db-1']
    ms2i_values = input_data['MS2i']
    for index, row in index_data.iterrows():
        ms2mz_NL2 = row['Precursor m/z-NL-2']
        ms2mz_NL4 = row['Precursor m/z-NL-4']
        ms2mz_NL6 = row['Precursor m/z-NL-6']
        ms2mz_NL8 = row['Precursor m/z-NL-8']
        ms2mz_NL10 = row['Precursor m/z-NL-10']
        ms2mz_NL12 = row['Precursor m/z-NL-12']
        ms2mz_NLdb = row['Precursor m/z-NL-db-1']
        matched_input_data_NL2 = input_data[(input_data['MS2mz'] - ms2mz_NL2).abs() < 0.05]
        matched_input_data_NL4 = input_data[(input_data['MS2mz'] - ms2mz_NL4).abs() < 0.05]
        matched_input_data_NL6 = input_data[(input_data['MS2mz'] - ms2mz_NL6).abs() < 0.05]
        matched_input_data_NL8 = input_data[(input_data['MS2mz'] - ms2mz_NL8).abs() < 0.05]
        matched_input_data_NL10 = input_data[(input_data['MS2mz'] - ms2mz_NL10).abs() < 0.05]
        matched_input_data_NL12 = input_data[(input_data['MS2mz'] - ms2mz_NL12).abs() < 0.05]
        matched_input_data_NLdb = input_data[(input_data['MS2mz'] - ms2mz_NLdb).abs() < 0.05]
        if not matched_input_data_NL2.empty:
            if matched_input_data_NLdb.empty:
                nl_db_intensity = 0
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                max_nl2_index = nl2_intensity.idxmax()
                nl2_intensity = nl2_intensity[max_nl2_index]
            else:
                nl2_intensity = ms2i_values[matched_input_data_NL2.index]
                nl_db_intensity = ms2i_values[matched_input_data_NLdb.index]
                max_nl2_index = nl2_intensity.idxmax()
                min_nl_db_index = (ms2mz_NLdb - input_data.loc[matched_input_data_NLdb.index, 'MS2mz']).abs().idxmin()
                nl2_intensity = nl2_intensity[max_nl2_index]
                nl_db_intensity = nl_db_intensity[min_nl_db_index]
            if (nl2_intensity > nl_db_intensity).all():
                input_data.loc[matched_input_data_NL2.index, 'double bond-NL-2'] = row['double bond position']
                input_data.loc[matched_input_data_NL2.index, 'NL-2 Matched'] = row['Precursor m/z-NL-2']
                input_data.loc[matched_input_data_NL4.index, 'NL-4 Matched'] = row['Precursor m/z-NL-4']
                input_data.loc[matched_input_data_NL6.index, 'NL-6 Matched'] = row['Precursor m/z-NL-6']
                input_data.loc[matched_input_data_NL8.index, 'NL-8 Matched'] = row['Precursor m/z-NL-8']
                input_data.loc[matched_input_data_NL10.index, 'NL-10 Matched'] = row['Precursor m/z-NL-10']
                input_data.loc[matched_input_data_NL12.index, 'NL-12 Matched'] = row['Precursor m/z-NL-12']
    columns_to_check = ['double bond-NL-2', 'NL-2 Matched', 'NL-4 Matched', 'NL-6 Matched', 'NL-8 Matched', 'NL-10 Matched', 'NL-12 Matched']
    for column in columns_to_check:
        if column in input_data.columns:
            input_data['double bond-NL-2'] = input_data['double bond-NL-2'].mask(input_data['double bond-NL-2'].duplicated(), '')
            input_data['NL-2 Matched'] = input_data['NL-2 Matched'].mask(input_data['NL-2 Matched'].duplicated(), '')
            input_data['NL-4 Matched'] = input_data['NL-4 Matched'].mask(input_data['NL-4 Matched'].duplicated(), '')
            input_data['NL-6 Matched'] = input_data['NL-6 Matched'].mask(input_data['NL-6 Matched'].duplicated(), '')
            input_data['NL-8 Matched'] = input_data['NL-8 Matched'].mask(input_data['NL-8 Matched'].duplicated(), '')
            input_data['NL-10 Matched'] = input_data['NL-10 Matched'].mask(input_data['NL-10 Matched'].duplicated(), '')
            input_data['NL-12 Matched'] = input_data['NL-12 Matched'].mask(input_data['NL-12 Matched'].duplicated(), '')
    return input_data
def process_excel(input_folder, index_file, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    index_data = {}
    for i in range(1, 7):
        sheet_name = f'DU-{i}'
        index_data[sheet_name] = pd.read_excel(index_file, sheet_name=sheet_name)
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            input_file = os.path.join(input_folder, filename)
            output_file = os.path.join(output_folder, filename)
            with pd.ExcelFile(input_file) as xls:
                with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                    for sheet_name in xls.sheet_names:
                        input_data = pd.read_excel(input_file, sheet_name=sheet_name)
                        _3_column = 'sn_3_chain_f' in input_data.columns
                        _2_column = 'sn_2_chain_b' in input_data.columns
                        if 'sn_1_chain_d' not in input_data.columns:
                            input_data.to_excel(writer, index=False, sheet_name=sheet_name)
                            continue
                        processed_data = pd.DataFrame()
                        for i in range(1, 7):
                            if (i in input_data['sn_1_chain_d'].values) or (_2_column and i in input_data['sn_2_chain_b'].values) or (_3_column and i in input_data['sn_3_chain_f'].values):
                                if f'process_excel_DU{i}' in globals():
                                    process_func = globals()[f'process_excel_DU{i}']
                                    processed_data = pd.concat([processed_data, process_func(input_data.copy(), index_data[f'DU-{i}'])], ignore_index=True)
                        processed_data.to_excel(writer, index=False, sheet_name=sheet_name)


_warned_messages = set()

def warn_once(message: str):
    """只在第一次看到 message 时打印一次"""
    if message not in _warned_messages:
        print(f"Warning: {message}")
        _warned_messages.add(message)


nl_columns = [
    "NL-1 Matched", "NL-2 Matched", "NL-4 Matched",
    "NL-6 Matched", "NL-8 Matched", "NL-10 Matched", "NL-12 Matched"
]
sn_columns = [
    ("sn-1-C-diagnostic ions", "sn-1-diagnostic ions", "sn-1-com"),
    ("sn-2-C-diagnostic ions", "sn-2-diagnostic ions", "sn-2-com"),
    ("sn-3-C-diagnostic ions", "sn-3-diagnostic ions", "sn-3-com")
]
required_columns = [
    "ID", "Subclass", "Precursor RT", "Precursor m/z", "TheoMz",
    "Formula", "ADDUCT_TYPE", "sn-composition", "double bond-NL-2",
    "Annotation level", "sn-diagnostic ions", "C=C diagnostic ions"
]

def process_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if 'Subclass' not in df.columns:
        warn_once("'Subclass' column not found!")
        return None  

    subclass_targets = ['LPC', 'LPC-O', 'LPC-P', 'LPE', 'LPE-O', 'LPE-P']
    first = df['Subclass'].iloc[0]
    if first in subclass_targets:
        annotation_level = 'level 1' if 'double bond-NL-2' in df.columns else 'level 2'
    else:
        if 'sn-2-com' in df.columns and 'double bond-NL-2' in df.columns:
            annotation_level = 'level 1'
        elif 'sn-2-com' in df.columns:
            annotation_level = 'level 2'
        else:
            annotation_level = 'level 3'

    df.loc[0, 'Annotation level'] = annotation_level
    return df

def process_excel_step2(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return None

    if 'MS2mz' not in df.columns or 'MS2i' not in df.columns:
        warn_once("'MS2mz' or 'MS2i' column not found!")
        return None  # 跳过

    
    df['MS2mz'] = df['MS2mz'].astype(object)
    df['MS2i']  = df['MS2i'].astype(object)

    
    columns_to_drop = ['sn_2_chain_b', 'sn_1_chain_d', 'sn_3_chain_f']
    df.drop(columns=[c for c in columns_to_drop if c in df.columns], inplace=True)

    
    df.at[0, 'MS2mz'] = ','.join(map(str, df['MS2mz'].dropna().tolist()))
    df.at[0, 'MS2i']  = ','.join(map(str, df['MS2i'].dropna().tolist()))
    df.loc[1:, ['MS2mz', 'MS2i']] = ''

    # C=C diagnostic ions
    def make_cc(row):
        return ";".join(str(row[c]) for c in nl_columns if c in row and pd.notna(row[c]))
    df['C=C diagnostic ions'] = df.apply(make_cc, axis=1)

    # sn-diagnostic ions
    def make_sn(row):
        parts = []
        for cols in sn_columns:
            valid = [str(row[c]) for c in cols if c in row and pd.notna(row[c])]
            if valid:
                parts.append(",".join(valid))
        return ";".join(parts)
    df['sn-diagnostic ions'] = df.apply(make_sn, axis=1)

    
    for col in required_columns:
        if col not in df.columns:
            warn_once(f"Missing column '{col}', filling with None.")
            df[col] = None

    
    df = df[required_columns]

    
    if 'ADDUCT_TYPE' in df.columns:
        df.loc[1:, 'ADDUCT_TYPE'] = ''

    return df

def process_and_merge_files(input_folder: str) -> dict:
    formatted = {}
    for fname in os.listdir(input_folder):
        if not fname.endswith(".xlsx"):
            continue
        path = os.path.join(input_folder, fname)
        xls = pd.ExcelFile(path)
        sheet_out = {}
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name)
            ps = process_sheet(df)
            pe = process_excel_step2(ps)
            if pe is not None:
                sheet_out[name] = pe
        if sheet_out:
            formatted[fname] = sheet_out
    return formatted

def join_unique(col: str, df: pd.DataFrame) -> str:
    if col not in df.columns:
        return ""
    vals = df[col].dropna().unique()
    vals = [v for v in vals if v != ""]
    return "/".join(map(str, vals))

def simplify_and_integrate_data(merged: dict) -> dict:
    out = {}
    for fname, sheets in merged.items():
        key = fname.rsplit('_', 1)[0]
        for name, df in sheets.items():
            sn_pos  = join_unique('sn-composition', df) or "No data"
            cc_pos  = join_unique('double bond-NL-2', df) or "No data"
            sn_diag = join_unique('sn-diagnostic ions', df) or "No data"
            cc_diag = join_unique('C=C diagnostic ions', df) or "No data"

            df.at[0, 'sn-position'] = sn_pos
            df.at[0, 'C=C position'] = cc_pos
            df.at[0, 'sn-diagnostic ions'] = sn_diag
            df.at[0, 'C=C diagnostic ions'] = cc_diag

            df = df.head(1)
            out.setdefault(key, []).append(df)
    return out

def save_integrated_data(dataframes: dict, output_folder: str):
    os.makedirs(output_folder, exist_ok=True)
    for key, dfs in dataframes.items():
        out_path = os.path.join(output_folder, f"{key}.xlsx")
        with pd.ExcelWriter(out_path, engine='xlsxwriter') as writer:
            for idx, df in enumerate(dfs, start=1):
                df2 = df.drop(columns=['sn-composition', 'double bond-NL-2'], errors='ignore')
                df2.to_excel(writer, sheet_name=f"Sheet_{idx}", index=False)

def combine_sheets(input_folder: str, output_folder: str):
    os.makedirs(output_folder, exist_ok=True)
    for fname in os.listdir(input_folder):
        if not fname.endswith('.xlsx'):
            continue
        path = os.path.join(input_folder, fname)
        xls = pd.ExcelFile(path)
        combined = pd.DataFrame()
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name)
            combined = pd.concat([combined, df], ignore_index=True)

        wb = Workbook()
        ws = wb.active
        for r, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        wb.save(os.path.join(output_folder, fname))


def cleanup_intermediate_folders():
    """删除所有中间文件夹"""
    intermediate_folders = [
        "output-step1",
        "output-step2",
        "output-step3",
        "output-step4",
        "output-step5"
    ]

    for folder in intermediate_folders:
        if os.path.exists(folder):
            try:
                shutil.rmtree(folder)
            except Exception as e:
                print(f"删除文件夹 {folder} 时出错: {e}")


if __name__ == "__main__":
    output_folders = [
        "output-step1",
        "output-step2",
        "output-step3",
        "output-step4",
        "output-step5",
        "final_results"
    ]

    
    for folder in output_folders:
        os.makedirs(folder, exist_ok=True)

    try:
        input_folder1 = "input-mzml"
        split_folder1 = "output-step1"
        main(input_folder1, split_folder1)

        input_folder2 = "output-step1"
        output_folder2 = "output-step2"
        process_files_and_profile(input_folder2, output_folder2)

        process_all(output_folder2, 'index-step3.xlsx', 'output-step3', verbose=False)

        input_folder4 = 'output-step3'
        index_file4 = 'index-step4.xlsx'
        output_folder4 = 'output-step4'
        process_original_data(input_folder4, index_file4, output_folder4)

        process_excel('output-step4', 'index-step5.xlsx', 'output-step5')

        input_folder6 = "output-step5"
        temp_folder6 = "final_results"
        merged = process_and_merge_files(input_folder6)
        integrated = simplify_and_integrate_data(merged)
        save_integrated_data(integrated, temp_folder6)
        combine_sheets(temp_folder6, temp_folder6)

        print("Processing completed.")

    finally:
        

        cleanup_intermediate_folders()
